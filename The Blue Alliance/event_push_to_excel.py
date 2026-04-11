import json
import os
import json
import os

from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.formatting.rule import FormulaRule, ColorScaleRule
from openpyxl.utils import get_column_letter
from dotenv import load_dotenv 

# Define the status message function.
def status(message):
    print(f"{datetime.now()}: {message}")


# Load the .env file and all environment variables.
load_dotenv()

# Retrieve values from .env.
tbaEventYear = os.getenv("TBA_EVENT_YEAR")
if tbaEventYear is None:
    raise ValueError("TBA_EVENT_YEAR is not set")

tbaEventKey = os.getenv("TBA_EVENT_KEY")
if tbaEventKey is None:
    raise ValueError("TBA_EVENT_KEY is not set")

status(f"Processing {tbaEventKey}")


# Define folder paths.
current_directory = os.path.dirname(os.path.abspath(__file__))
target_event_directory = os.path.join(current_directory, "Game Years", tbaEventYear, tbaEventKey)
template_directory = os.path.join(current_directory, "Templates")
os.makedirs(target_event_directory, exist_ok=True)


# Set various variables we can use throughout the script.
template_file_name = os.path.join(template_directory, "scouting_template.xlsx")
ouput_file_name = os.path.join(target_event_directory, f"{tbaEventKey}.xlsx")


# Set constants we can use.
font_header    = Font(bold=True,  size=12, color="FFFFFF")
font_data      = Font(bold=False, size=12, color="000000")
font_link      = Font(bold=False, size=12, color="000000", underline='single')

align_center   = Alignment(horizontal="center", vertical="center")
align_vertical = Alignment(horizontal="left",   vertical="bottom", text_rotation=90) 

fill_default   = PatternFill(start_color="595959", end_color="595959", fill_type="solid") 
fill_blue      = PatternFill(start_color="2471A3", end_color="2471A3", fill_type="solid")   
fill_red       = PatternFill(start_color="CB4335", end_color="CB4335", fill_type="solid")   
fill_green     = PatternFill(start_color="229954", end_color="229954", fill_type="solid")   
fill_orange    = PatternFill(start_color="D68910", end_color="D68910", fill_type="solid")  

color_scale_rule = ColorScaleRule(
    start_type="min", start_color="CB4335",                     # Red for minimum value
    mid_type="percentile", mid_value=50, mid_color="D68910",    # Yellow for mid value
    end_type="max", end_color="229954"                          # Green for maximum value
)


format_decimal = '#,##0.00;-[Red]#,##0.00;"-"'
format_comma = '#,##0;-[Red]#,##0;"-"'
format_percent = '0.0%'


# Load the workbook and save it as the destination workbook.
status(f"Preparing the spreadsheet: {ouput_file_name}")
wb = load_workbook(template_file_name)
wb.save(ouput_file_name)

# Workaround for the inability to do something like this:
# row[counter += 1]
def incrementer(start):
    while True:
        start += 1
        yield start


# Define the function that can freeze panes and set auto filter.
def apply_view(ws, freeze_range, auto_filter_range):

    if not freeze_range is None:
        ws.freeze_panes =freeze_range

    if not auto_filter_range is None:
        ws.auto_filter.ref = auto_filter_range     


# Accepts a sheet and a list of format definitions and applies them to the sheet.
# This allows us to apply a variable number of formats to cell ranges.
def apply_formats(ws, formats):
    for format in formats:
        # Extract the range definition and the formats to apply.
        range = format["range"]
        fill = format.get("fill")
        font = format.get("font")
        alignment = format.get("alignment")
        number_format = format.get("number_format")

        # Enumerate over the cells in the range.
        for cell_row in ws[range]:

            # Apply formats to each individual cell.
            for cell in cell_row:
                if not fill is None:
                    cell.fill = fill
                if not font is None:
                    cell.font = font
                if not alignment is None:
                    cell.alignment = alignment
                if not number_format is None:
                    cell.number_format = number_format


# Define the the function that reads and pushes the Event.
def prepate_sheet_event():
    status("Pushing Event data...")

    # Read the JSON data from the file.
    with open(os.path.join(target_event_directory, f"{tbaEventKey}.json"), "r") as f:
        data = json.load(f)

    # Load the Event sheet.
    ws = wb["Event"]

    # Write the Event data to the sheet.
    ws["B1"] = data["key"]
    ws["B2"] = data["event_code"]
    ws["B3"] = data["name"]
    ws["B4"] = data["location_name"]
    ws["B5"] = data["start_date"]

    # Apply formatting to the sheet.
    apply_formats(ws, [
        { "range": f"B1:B5", "font": font_data },
    ])


# Define the the function that reads and pushes the Teams.
def prepare_sheet_teams():
    status("Pushing Team data...")

    # Read the JSON data from the file.
    with open(os.path.join(target_event_directory, f"{tbaEventKey}.teams.json"), "r") as f:
        data = json.load(f)
        data = sorted(data, key=lambda x: x["team_number"])

    # Load the Teams sheet.
    ws = wb["Teams"]

    # Write the data to the sheet.
    for index, row in enumerate(data):
        row_num = index + 2
        team_number = row["team_number"]
        ws.cell(row=row_num, column=1, value=team_number)
        ws.cell(row=row_num, column=2, value=row["nickname"])
        ws.cell(row=row_num, column=3, value=row["school_name"])

        cell = ws.cell(row=row_num, column=4)
        cell.value = "Go to TBA" # type: ignore
        cell.hyperlink = f"https://www.thebluealliance.com/team/{team_number}" # type: ignore
        cell.style = "Hyperlink"
        cell.alignment = Alignment(horizontal='center')

        cell = ws.cell(row=row_num, column=5)
        cell.value = "Go to FRC" # type: ignore
        cell.hyperlink = f"https://frc-events.firstinspires.org/team/{team_number}" # type: ignore
        cell.style = "Hyperlink"
        cell.alignment = Alignment(horizontal='center')

        cell = ws.cell(row=row_num, column=6)
        cell.value = "Go to StatBotics" # type: ignore
        cell.hyperlink = f"https://www.statbotics.io/team/{team_number}/{tbaEventYear}" # type: ignore
        cell.style = "Hyperlink"                
        cell.alignment = Alignment(horizontal='center')
  
    # Apply formatting to the sheet.
    apply_formats(ws, [
        { "range": f"A2:C{str(len(data) + 1)}", "font": font_data },
    ])


# Define the the function that reads and pushes the Matches.
def prepare_sheet_matches():
    status("Pushing Match data...")

    # Set the conditional formatting rules.
    fill_scouted = PatternFill(start_color="C8D6A1", end_color="C8D6A1", fill_type="solid")
    fill_duplicate_match = PatternFill(start_color="FFFF54", end_color="FFFF54", fill_type="solid")
    fill_wrong_team = PatternFill(start_color="F5C242", end_color="F5C242", fill_type="solid")

    # Read the JSON data from the file.
    with open(os.path.join(target_event_directory, f"{tbaEventKey}.matches.json"), "r") as f:
        data = json.load(f)
        data = [row for row in data if row["comp_level"] == "qm"]
        data = sorted(data, key=lambda x: x["match_number"])

    # Load the Teams sheet.
    ws = wb["Matches"]

    # Write the data to the sheet.
    for index, row in enumerate(data):
        row_num = index + 2
        ws.cell(row=row_num, column=1, value=row["match_number"])
        ws.cell(row=row_num, column=2, value=datetime.fromtimestamp(row["time"]).strftime("%Y-%m-%d %H:%M:%S"))
        ws.cell(row=row_num, column=3, value=int(row["alliances"]["blue"]["team_keys"][0].replace("frc", "")))
        ws.cell(row=row_num, column=4, value=int(row["alliances"]["blue"]["team_keys"][1].replace("frc", "")))
        ws.cell(row=row_num, column=5, value=int(row["alliances"]["blue"]["team_keys"][2].replace("frc", "")))
        ws.cell(row=row_num, column=6, value=int(row["alliances"]["red"]["team_keys"][0].replace("frc", "")))
        ws.cell(row=row_num, column=7, value=int(row["alliances"]["red"]["team_keys"][1].replace("frc", "")))
        ws.cell(row=row_num, column=8, value=int(row["alliances"]["red"]["team_keys"][2].replace("frc", "")))

        # Add conditional formatting that cannot be applied to a range.
        for col in ["C", "D", "E", "F", "G", "H"]:
            scouted_wrong_team = FormulaRule(formula=[f"=IF(VLOOKUP(CONCATENATE($A{row_num}, \".\", ${col}$1),MatchScoutingData!$A$2:$E$1000, 5, FALSE) = ${col}{row_num}, FALSE, TRUE)"], fill=fill_wrong_team)
            ws.conditional_formatting.add(f"{col}{row_num}", scouted_wrong_team)

  
    # Apply formatting to the sheet.
    apply_formats(ws, [
        { "range": f"A1:B1", "font": font_header ,"fill": fill_default, "alignment": align_center },
        { "range": f"C1:E1", "font": font_header ,"fill": fill_blue, "alignment": align_center },
        { "range": f"F1:H1", "font": font_header ,"fill": fill_red, "alignment": align_center },
        { "range": f"A2:H{str(len(data) + 1)}", "font": font_data, "alignment": align_center },
    ])

    # Add conditional formatting rules that can be applied to a range.
    # =COUNTIF(Where do you want to look?, What do you want to look for?)
    rule_scouted = FormulaRule(formula=["=COUNTIF(MatchScoutingData!$A$2:$A$1000, CONCATENATE($A2,\".\",C$1)) = 1"], fill=fill_scouted)
    rule_scouted_multiple = FormulaRule(formula=["=COUNTIF(MatchScoutingData!$A$2:$A$1000, CONCATENATE($A2,\".\",C$1)) > 1"], fill=fill_duplicate_match)
    ws.conditional_formatting.add("C2:H1000", rule_scouted)
    ws.conditional_formatting.add("C2:H1000", rule_scouted_multiple)

    # Provide a key so it is wasy to understand conditional formatting.
    ws[f"J2"] = "Match has not yet been scouted"
    ws[f"J3"] = "Match has been scouted"
    ws[f"J4"] = "Match was scouted more than once"
    ws[f"J5"] = "Wrong team was scouted"

    # Apply formats to the key.
    apply_formats(ws, [
        { "range": f"J3:J3", "fill": fill_scouted },
        { "range": f"J4:J4", "fill": fill_duplicate_match },
        { "range": f"J5:J5", "fill": fill_wrong_team  },
    ])    


# Deine the function that prepares the Team Scores sheet.
def prepare_sheet_team_scores():
    status("Prepating Team Scores sheet...")

    # Read the JSON data from the file.
    with open(os.path.join(target_event_directory, f"{tbaEventKey}.teams.json"), "r") as f:
        data = json.load(f)
        data = sorted(data, key=lambda x: x["team_number"])

    # Capture the extents of the sheet and data.
    start_row = 3
    record_count = len(data)
    end_row = start_row + record_count - 1

    # Open the Teams sheet.
    ws = wb["Team Scores"]

    # Write the team numbers to the sheet.
    for index, row in enumerate(data):
        row_num = index + start_row
        ws.cell(row=row_num, column=1, value=row["team_number"])  

    # Apply the formulas.
    for row in range(start_row, start_row + record_count):
        # General
        ws[f"B{row}"] = f"=COUNTIF(MatchScoutingData!E$2:E$1000, $A{row})"      # Matches Scouted

        # Generate formulas to account for all the source columns.
        dest_start = 3   # Column C
        dest_end   = 23  # Column X
        src_start = 6    # Column F

        for i, col_idx in enumerate(range(dest_start, dest_end + 1)):
            target_col = get_column_letter(col_idx)
            source_col = get_column_letter(src_start + i)

            ws[f"{target_col}{row}"] = (
                f"=SUMIF(MatchScoutingData!$E$2:$E$1000, "
                f"$A{row}, "
                f"MatchScoutingData!{source_col}$2:{source_col}$1000)"
            )

        # Generate formula for new columns.
        ws[f"X{row}"]  = f"=SUM(U{row}:W{row})"     # Total Ranking Points
        ws[f"Y{row}"]  = f"=SUM(C{row}:E{row})"     # Auto
        ws[f"Z{row}"]  = f"=SUM(F{row}:O{row})"     # Teleop
        ws[f"AA{row}"] = f"=SUM(P{row}:R{row})"     # Endgame
        ws[f"AB{row}"] = f"=SUM(S{row}:X{row})"     # Post Match
        ws[f"AC{row}"] = f"=SUM(Y{row}:AB{row})"    # Complete

  
    # Apply formatting to the sheet.
    apply_formats(ws, [
        # Team / Scouted Count
        { "range": f"A{start_row}:B{end_row}", "font": font_header, "fill": fill_default },

        # Format the data cells.
        { "range": f"C{start_row}:Z{end_row}", "font": font_data, "number_format": format_comma },        
    ])


def prepare_sheet_team_summary():
    status("Prepating Team Summary Weightings sheet...")

    # Read the JSON data from the file.
    with open(os.path.join(target_event_directory, f"{tbaEventKey}.teams.json"), "r") as f:
        data = json.load(f)
        data = sorted(data, key=lambda x: x["team_number"])

    # Capture the extents of the sheet and data.
    record_count = len(data)
    start_row = 4
    end_row = start_row + record_count - 1

    # Open the Teams sheet.
    ws = wb["Team Summary"]

    # Write the team numbers to the sheet.
    for index, row in enumerate(data):
        row_num = index + start_row
        ws.cell(row=row_num, column=1, value=row["team_number"])

    # Apply the formulas.
    for row in range(start_row, start_row + record_count):
        # General
        ws[f"B{row}"] = f"=SUM(C{row}:AD{row})"

        # Generate formulas to account for all the source columns.
        dest_start = 3   # Column C
        dest_end   = 30  # Column AD
        src_start  = 2   # Column B

        for i, col_idx in enumerate(range(dest_start, dest_end + 1)):
            dest_col = get_column_letter(col_idx)
            src_col = get_column_letter(src_start + i)

            ws[f"{dest_col}{row}"] = (
                f"=PERCENTRANK('Team Scores'!${src_col}$3:${src_col}$1000, "
                f"'Team Scores'!{src_col}{row - 1}, 3) * {dest_col}$2"
            )

    # Apply the formats to the cells.
    apply_formats(ws, [
        # Team / Scouted Count
        { "range": f"A{start_row}:B{end_row}", "font": font_header, "fill": fill_default },

        # All data cells.
        #{ "range": f"B{start_row}:B{end_row}", "number_format": format_comma },
        { "range": f"C{start_row}:AD{end_row}", "number_format": format_percent },
    ])

    # Build and apply the conditional formatting rules to produce a heatmap for percentiles.
    ws.conditional_formatting.add(f"C{start_row}:AD{end_row}", color_scale_rule)


def prepare_sheet_tba_opr():
    status("Prepating The Blue Alliance OPR sheet...")

    # Read the JSON data from the file.
    with open(os.path.join(target_event_directory, f"{tbaEventKey}.teams.json"), "r") as f:
        data = json.load(f)
        data = sorted(data, key=lambda x: x["team_number"])

    # Capture the extents of the sheet and data.
    record_count = len(data)
    start_row = 2
    end_row = start_row + record_count - 1

    # Open the Teams sheet.
    ws = wb["TBA OPR"]

    # Write the team numbers to the sheet.
    for index, row in enumerate(data):
        row_num = index + start_row
        ws.cell(row=row_num, column=1, value=row["team_number"])

    # Apply the formulas.
    for row in range(start_row, start_row + record_count):
        # Generate formulas to account for all the source columns.
        dest_start = 2   # Column B
        dest_end   = 27  # Column AA

        for i, col_idx in enumerate(range(dest_start, dest_end + 1)):
            dest_col = get_column_letter(col_idx)

            # ws[f"{dest_col}{row}"] = (
            #     f"=PERCENTRANK(${dest_col}$2:${dest_col}$1000, "
            #     f"'{dest_col}{row}, 3)"
            # )

    # Apply the formats to the cells.
    apply_formats(ws, [
        # Team / Scouted Count
        { "range": f"A{start_row}:B{end_row}", "font": font_header, "fill": fill_default },

        # All data cells.
        { "range": f"B{start_row}:AA{end_row}", "number_format": format_decimal },
    ])

    # Build and apply the conditional formatting rules to produce a heatmap for percentiles.
    ws.conditional_formatting.add(f"B{start_row}:AD{end_row}", color_scale_rule)



# Push the data into the spreadsheet.
prepate_sheet_event()
prepare_sheet_teams()
prepare_sheet_matches()
prepare_sheet_team_scores()
prepare_sheet_team_summary()
prepare_sheet_tba_opr()

# And finally, save the spreadsheet.
status(f"Saving to {ouput_file_name}...")
wb.save(ouput_file_name)

