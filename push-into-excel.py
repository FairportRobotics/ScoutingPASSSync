import json
import os
import json
import os

from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import FormulaRule, ColorScaleRule
from dotenv import load_dotenv 

# Load the .env file and all environment variables.
load_dotenv()

# Define the status message function.
def status(message):
    print(f"{datetime.now()}: {message}")

# Retrieve values from .env.
tbaEventKey: str = os.getenv("TBA_EVENT_KEY")   # From FIRST/The Blue Alliance

# Perform validations.
if tbaEventKey == "":
    status("No event key provided.")
    exit()   

if not os.path.exists("push-into-excel-template.xlsx"):
    status("push-into-excel.xlsx cannot be found")
    exit()   

# Set various variables we can use throughout the script.
ouput_file_name=f"excel_data/Match Scouting Results.xlsx"

font_header = Font(bold=True, size=12, color="FFFFFF")
font_data = Font(bold=False, size=12, color="000000")

align_center = Alignment(horizontal="center", vertical="center")
align_vertical = Alignment(horizontal="left", vertical="bottom", text_rotation=90) 

fill_default = PatternFill(start_color="595959", end_color="595959", fill_type="solid") 
fill_blue = PatternFill(start_color="2471a3", end_color="2471a3", fill_type="solid")   
fill_red = PatternFill(start_color="cb4335", end_color="cb4335", fill_type="solid")   
fill_green = PatternFill(start_color="229954", end_color="229954", fill_type="solid")   
fill_orange = PatternFill(start_color="D68910", end_color="D68910", fill_type="solid")  

color_scale_rule = ColorScaleRule(
    start_type="min", start_color="cb4335",                     # Red for minimum value
    mid_type="percentile", mid_value=50, mid_color="D68910",    # Yellow for mid value
    end_type="max", end_color="229954"                          # Green for maximum value
)

format_comma = '#,##0;[Red]#,##0;"-"'
format_percent = '0.0%'

# Load the workbook and save it as the destination workbook.
status(f"Preparing the spreadsheet: {ouput_file_name}")
wb = load_workbook("push-into-excel-template.xlsx")
wb.save(ouput_file_name)


# Workaround for the inability to do something like this:
# row[counter += 1]
def incrementer(start):
    while True:
        start += 1
        yield start


# Define the function that can freeze panes and set auto filter.
def apply_view(ws, freeze_range, auto_filter_range):

    if not freeze_range is None:
        ws.freeze_panes =freeze_range

    if not auto_filter_range is None:
        ws.auto_filter.ref = auto_filter_range     


# Accepts a sheet and a list of format definitions and applies them to the sheet.
# This allows us to apply a variable number of formats to cell ranges.
def apply_formats(ws, formats):
    for format in formats:
        # Extract the range definition and the formats to apply.
        range = format["range"]
        fill = format.get("fill")
        font = format.get("font")
        alignment = format.get("alignment")
        number_format = format.get("number_format")

        # Enumerate over the cells in the range.
        for cell_row in ws[range]:

            # Apply formats to each individual cell.
            for cell in cell_row:
                if not fill is None:
                    cell.fill = fill
                if not font is None:
                    cell.font = font
                if not alignment is None:
                    cell.alignment = alignment
                if not number_format is None:
                    cell.number_format = number_format


# Define the the function that reads and pushes the Event.
def prepate_sheet_event():
    status("Pushing Event data...")

    # Read the JSON data from the file.
    file=f"tba_data/{tbaEventKey}.json"
    with open(file, "r") as f:
        data = json.load(f)

    # Load the Event sheet.
    ws = wb["Event"]

    # Write the Event data to the sheet.
    ws["B1"] = data["key"]
    ws["B2"] = data["event_code"]
    ws["B3"] = data["name"]
    ws["B4"] = data["location_name"]
    ws["B5"] = data["start_date"]

    # Apply formatting to the sheet.
    apply_formats(ws, [
        { "range": f"B1:B5", "font": font_data },
    ])


# Define the the function that reads and pushes the Teams.
def prepare_sheet_teams():
    status("Pushing Team data...")

    # Read the JSON data from the file.
    file=f"tba_data/{tbaEventKey}.teams.json"
    with open(file, "r") as f:
        data = json.load(f)
        data = sorted(data, key=lambda x: x["team_number"])

    # Load the Teams sheet.
    ws = wb["Teams"]

    # Write the data to the sheet.
    for index, row in enumerate(data):
        row_num = index + 2
        ws.cell(row=row_num, column=1, value=row["team_number"])
        ws.cell(row=row_num, column=2, value=row["nickname"])
        ws.cell(row=row_num, column=3, value=row["school_name"])
        ws.cell(row=row_num, column=4, value=row["rookie_year"])
        ws.cell(row=row_num, column=5, value=row["key"])
  
    # Apply formatting to the sheet.
    apply_formats(ws, [
        { "range": f"A2:A{str(len(data) + 1)}", "font": font_data },
    ])


# Define the the function that reads and pushes the Matches.
def prepare_sheet_matches():
    status("Pushing Match data...")
    file=f"tba_data/{tbaEventKey}.matches.json"

    # Read the JSON data from the file.
    with open(file, "r") as f:
        data = json.load(f)
        data = [row for row in data if row["comp_level"] == "qm"]
        data = sorted(data, key=lambda x: x["match_number"])

    # Load the Teams sheet.
    ws = wb["Matches"]

    # Write the data to the sheet.
    for index, row in enumerate(data):
        row_num = index + 2
        ws.cell(row=row_num, column=1, value=row["match_number"])
        ws.cell(row=row_num, column=2, value=datetime.fromtimestamp(row["time"]).strftime("%Y-%m-%d %H:%M:%S"))
        ws.cell(row=row_num, column=3, value=int(row["alliances"]["blue"]["team_keys"][0].replace("frc", "")))
        ws.cell(row=row_num, column=4, value=int(row["alliances"]["blue"]["team_keys"][1].replace("frc", "")))
        ws.cell(row=row_num, column=5, value=int(row["alliances"]["blue"]["team_keys"][2].replace("frc", "")))
        ws.cell(row=row_num, column=6, value=int(row["alliances"]["red"]["team_keys"][0].replace("frc", "")))
        ws.cell(row=row_num, column=7, value=int(row["alliances"]["red"]["team_keys"][1].replace("frc", "")))
        ws.cell(row=row_num, column=8, value=int(row["alliances"]["red"]["team_keys"][2].replace("frc", "")))
  
    # Apply formatting to the sheet.
    apply_formats(ws, [
        { "range": f"A1:B1", "font": font_header ,"fill": fill_default, "alignment": align_center },
        { "range": f"C1:E1", "font": font_header ,"fill": fill_blue, "alignment": align_center },
        { "range": f"F1:H1", "font": font_header ,"fill": fill_red, "alignment": align_center },
        { "range": f"A2:H{str(len(data) + 1)}", "font": font_data, "alignment": align_center },
    ])

    # Set the conditional formatting rules.
    fill_scouted = PatternFill(start_color="C8D6A1", end_color="C8D6A1", fill_type="solid")
    fill_scouted_multiple = PatternFill(start_color="D09996", end_color="D09996", fill_type="solid")

    rule_scouted = FormulaRule(formula=["=COUNTIF(MatchScoutingData!$A$2:$A$1000,CONCATENATE($A2,\".\",C$1)) = 1"], fill=fill_scouted)
    rule_scouted_multiple = FormulaRule(formula=["=COUNTIF(MatchScoutingData!$A$2:$A$1000,CONCATENATE($A2,\".\",C$1)) > 1"], fill=fill_scouted_multiple)

    ws.conditional_formatting.add("C2:H1000", rule_scouted)
    ws.conditional_formatting.add("C2:H1000", rule_scouted_multiple)


# Deine the function that prepares the Team Scores sheet.
def prepare_sheet_team_scores():
    status("Prepating Team Scores sheet...")

    # Read the JSON data from the file.
    file=f"tba_data/{tbaEventKey}.teams.json"
    with open(file, "r") as f:
        data = json.load(f)
        data = sorted(data, key=lambda x: x["team_number"])

    # Capture the extents of the sheet and data.
    start_row = 4
    record_count = len(data)
    last_row = start_row + record_count - 1

    # Open the Teams sheet.
    ws = wb["Team Scores"]

    # Write the team numbers to the sheet.
    for index, row in enumerate(data):
        row_num = index + start_row
        ws.cell(row=row_num, column=1, value=row["team_number"])  

    # Apply the formulas.
    for row in range(start_row, start_row + record_count):
        #General
        ws[f"B{row}"] = f"=COUNTIF(MatchScoutingData!G$2:G$1000, $A{row})"                                  # #Scouted

        # Autos
        ws[f"C{row}"] = f"=SUMIF(MatchScoutingData!$G$2:$G$1000, $A{row}, MatchScoutingData!J$2:J$1000)"    # # L1
        ws[f"D{row}"] = f"=SUMIF(MatchScoutingData!$G$2:$G$1000, $A{row}, MatchScoutingData!K$2:K$1000)"    # # L2
        ws[f"E{row}"] = f"=SUMIF(MatchScoutingData!$G$2:$G$1000, $A{row}, MatchScoutingData!L$2:L$1000)"    # # L3
        ws[f"F{row}"] = f"=SUMIF(MatchScoutingData!$G$2:$G$1000, $A{row}, MatchScoutingData!M$2:M$1000)"    # # L4
        ws[f"G{row}"] = f"=SUMIF(MatchScoutingData!$G$2:$G$1000, $A{row}, MatchScoutingData!N$2:N$1000)"    # # Proc
        ws[f"H{row}"] = f"=SUMIF(MatchScoutingData!$G$2:$G$1000, $A{row}, MatchScoutingData!O$2:O$1000)"    # # Net

        ws[f"I{row}"] = f"=C{row} * 3"    # P L1
        ws[f"J{row}"] = f"=D{row} * 4"    # P L2
        ws[f"K{row}"] = f"=E{row} * 6"    # P L3
        ws[f"L{row}"] = f"=F{row} * 7"    # P L4
        ws[f"M{row}"] = f"=G{row} * 6"    # P Proc
        ws[f"N{row}"] = f"=H{row} * 4"    # P Net    

        ws[f"O{row}"] = f"=SUMIF(MatchScoutingData!$G$2:$G$1000, $A{row}, MatchScoutingData!I$2:I$1000) * 3"    # P Leave    
        ws[f"P{row}"] = f"=SUM(I{row}:O{row})"    # AP Total    
        ws[f"Q{row}"] = f"=IF(P{row} > 0, P{row}/B{row}, 0)"    # AP Avg/Mat

        # Teleop
        ws[f"R{row}"] = f"=SUMIF(MatchScoutingData!$G$2:$G$1000, $A{row}, MatchScoutingData!P$2:P$1000)"    # # L1
        ws[f"S{row}"] = f"=SUMIF(MatchScoutingData!$G$2:$G$1000, $A{row}, MatchScoutingData!Q$2:Q$1000)"    # # L2
        ws[f"T{row}"] = f"=SUMIF(MatchScoutingData!$G$2:$G$1000, $A{row}, MatchScoutingData!R$2:R$1000)"    # # L3
        ws[f"U{row}"] = f"=SUMIF(MatchScoutingData!$G$2:$G$1000, $A{row}, MatchScoutingData!S$2:S$1000)"    # # L4
        ws[f"V{row}"] = f"=SUMIF(MatchScoutingData!$G$2:$G$1000, $A{row}, MatchScoutingData!T$2:T$1000)"    # # Proc
        ws[f"W{row}"] = f"=SUMIF(MatchScoutingData!$G$2:$G$1000, $A{row}, MatchScoutingData!U$2:U$1000)"    # # Net

        ws[f"X{row}"]  = f"=R{row} * 2"    # P L1
        ws[f"Y{row}"]  = f"=S{row} * 3"    # P L2
        ws[f"Z{row}"]  = f"=T{row} * 4"    # P L3
        ws[f"AA{row}"] = f"=U{row} * 5"    # P L4
        ws[f"AB{row}"] = f"=V{row} * 6"    # P Proc
        ws[f"AC{row}"] = f"=W{row} * 4"    # P Net    

        ws[f"AD{row}"] = f"=SUM(X{row}:AC{row})"    # TP Total    
        ws[f"AE{row}"] = f"=IF(AD{row} > 0, AD{row}/B{row}, 0)"    # TP Avg/Mat        

        # Endgame
        ws[f"AF{row}"] = f"=COUNTIFS(MatchScoutingData!$G$2:$G$1000, $A{row}, MatchScoutingData!$W$2:$W$1000, \"p\")"    # # Barge P
        ws[f"AG{row}"] = f"=COUNTIFS(MatchScoutingData!$G$2:$G$1000, $A{row}, MatchScoutingData!$W$2:$W$1000, \"s\")"    # # Barge S
        ws[f"AH{row}"] = f"=COUNTIFS(MatchScoutingData!$G$2:$G$1000, $A{row}, MatchScoutingData!$W$2:$W$1000, \"d\")"    # # Barge D

        ws[f"AI{row}"] = f"=AF{row} * 2"     # P Barge P
        ws[f"AJ{row}"] = f"=AG{row} * 6"     # P Barge S
        ws[f"AK{row}"] = f"=AH{row} * 12"    # P Barge D

        ws[f"AL{row}"] = f"=SUM(AI{row}:AK{row})"               # EP Total    
        ws[f"AM{row}"] = f"=IF(AD{row} > 0, AL{row}/B{row}, 0)" # EP Avg/Mat     

        # Totals
        ws[f"AN{row}"] = f"=SUM(P{row}, AD{row}, AL{row})"    # P Total
        ws[f"AO{row}"] = f"=IF(AN{row} > 0, AN{row}/B{row}, 0)"    # Avg/Mat        

  
    # Apply formatting to the sheet.
    apply_formats(ws, [
        # Format the data cells.
        { "range": f"B{start_row}:AO{last_row}", "font": font_data, "number_format": format_comma },        

        # Format the fill color of the data rows.
        { "range": f"A{start_row}:A{last_row}", "font": font_header ,"fill": fill_default },
        { "range": f"P{start_row}:Q{last_row}", "font": font_header ,"fill": fill_blue },
        { "range": f"AD{start_row}:AE{last_row}", "font": font_header ,"fill": fill_red },
        { "range": f"AL{start_row}:AM{last_row}", "font": font_header ,"fill": fill_orange },
        { "range": f"AN{start_row}:AO{last_row}", "font": font_header ,"fill": fill_green },
    ])


# Define the function that prepares the Team Summary sheet.
def prepare_sheet_team_summary():
    status("Prepating Team Summary sheet...")

    # Read the JSON data from the file.
    file=f"tba_data/{tbaEventKey}.teams.json"
    with open(file, "r") as f:
        data = json.load(f)
        data = sorted(data, key=lambda x: x["team_number"])

    # Capture the extents of the sheet and data.
    record_count = len(data)
    start_row = 4
    end_row = start_row + record_count - 1
    source_start_row = 4
    source_end_row = source_start_row + record_count - 1        

    # Open the Teams sheet.
    ws = wb["Team Summary"]

    # Write the team numbers to the sheet.
    for index, row in enumerate(data):
        row_num = index + start_row
        ws.cell(row=row_num, column=1, value=row["team_number"])

    # Apply the formulas.
    for row in range(start_row, end_row + 1):

        # Points Scored
        ws[f"B{row}"] = f"=INDEX('Team Scores'!P${source_start_row}:P${source_end_row}, MATCH(A{row},'Team Scores'!A$4:A${source_end_row}, 0))"
        ws[f"C{row}"] = f"=INDEX('Team Scores'!Q${source_start_row}:Q${source_end_row}, MATCH(A{row},'Team Scores'!A$4:A${source_end_row}, 0))"

        ws[f"D{row}"] = f"=INDEX('Team Scores'!AD${source_start_row}:AD${source_end_row}, MATCH(A{row},'Team Scores'!A$4:A${source_end_row}, 0))"
        ws[f"E{row}"] = f"=INDEX('Team Scores'!AE${source_start_row}:AE${source_end_row}, MATCH(A{row},'Team Scores'!A$4:A${source_end_row}, 0))"

        ws[f"F{row}"] = f"=INDEX('Team Scores'!AL${source_start_row}:AL${source_end_row}, MATCH(A{row},'Team Scores'!A$4:A${source_end_row}, 0))"
        ws[f"G{row}"] = f"=INDEX('Team Scores'!AM${source_start_row}:AM${source_end_row}, MATCH(A{row},'Team Scores'!A$4:A${source_end_row}, 0))"

        ws[f"H{row}"] = f"=INDEX('Team Scores'!AN${source_start_row}:AN${source_end_row}, MATCH(A{row},'Team Scores'!A$4:A${source_end_row}, 0))"
        ws[f"I{row}"] = f"=INDEX('Team Scores'!AO${source_start_row}:AO${source_end_row}, MATCH(A{row},'Team Scores'!A$4:A${source_end_row}, 0))"

        # Percentile
        ws[f"J{row}"] = f"=PERCENTRANK(B$2:B${end_row}, B{row}, 3)"
        ws[f"K{row}"] = f"=PERCENTRANK(D$2:D${end_row}, D{row}, 3)"
        ws[f"L{row}"] = f"=PERCENTRANK(F$2:F${end_row}, F{row}, 3)"
        ws[f"M{row}"] = f"=PERCENTRANK(H$2:H${end_row}, H{row}, 3)"

    # Apply the formats to the cells.
    apply_formats(ws, [
        # Team
        { "range": f"A{start_row}:A{end_row}", "font": font_header, "fill": fill_default },

        # All data cells.
        { "range": f"B{start_row}:M{end_row}", "font": font_data, "number_format": format_comma },
        { "range": f"J{start_row}:M{end_row}", "number_format": format_percent },
        { "range": f"R{start_row}:M{end_row}", "number_format": format_percent },
    ])

    # Build and apply the conditional formatting rules to produce a heatmap for percentiles.
    ws.conditional_formatting.add(f"J{start_row}:M{end_row}", color_scale_rule)
    ws.conditional_formatting.add(f"R{start_row}:U{end_row}", color_scale_rule)


# Hello World
def prepare_sheet_pbi_team_summary():
    status("Prepating Power BI Team Summary sheet...")

    # Open the Teams sheet.
    ws = wb["PB Team Summary"]   

    ws["A2"] = "=SORT(UNIUE())"    


# Hello World
def prepare_sheet_pbi_scouter_summary():
    status("Prepating Power BI Scouter Summary sheet...")

    # Read the JSON data from the file.
    file=f"tba_data/{tbaEventKey}.teams.json"
    with open(file, "r") as f:
        data = json.load(f)
        data = sorted(data, key=lambda x: x["team_number"])

    # Capture the extents of the sheet and data.
    record_count = len(data)
    start_row = 2

    # Open the Teams sheet.
    ws = wb["PB Team Summary"]    

    # Apply the formulas.
    for row in range(start_row, start_row + record_count):
        ws[f"A{row}"] = f"='Team Summary'!A{row + 2}"
        ws[f"B{row}"] = f"='Team Summary'!B{row + 2}"
        ws[f"C{row}"] = f"='Team Summary'!C{row + 2}"
        ws[f"D{row}"] = f"='Team Summary'!D{row + 2}"
        ws[f"E{row}"] = f"='Team Summary'!E{row + 2}"
        ws[f"F{row}"] = f"='Team Summary'!F{row + 2}"
        ws[f"G{row}"] = f"='Team Summary'!G{row + 2}"


# Push the data into the spreadsheet.
prepate_sheet_event()
prepare_sheet_teams()
prepare_sheet_matches()
prepare_sheet_team_scores()
prepare_sheet_team_summary()

prepare_sheet_pbi_team_summary()
prepare_sheet_pbi_scouter_summary()


# And finally, save the spreadsheet.
wb.save(ouput_file_name)
